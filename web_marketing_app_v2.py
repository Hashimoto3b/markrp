import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ページ設定
st.set_page_config(page_title="Webマーケティング レポート自動化ツール", layout="centered")

st.title("📊 Webマーケティング自動レポート作成ツール（v2）")
st.markdown("**ようこそ！このアプリではマーケティングデータをアップロードするだけで、自動で集計・分析・改善提案を行います。**")

# 初めにやることガイド
with st.expander("🟢 はじめにやること（クリックで開閉）", expanded=True):
    st.markdown("""
1. **CSVまたはExcelファイル**をアップロードしてください（複数可）  
2. データに「日付」「売上」「クリック」「コンバージョン」「費用」などの列が含まれていると、各種KPI（LTV/ROAS/ROI/CPA）が自動算出されます  
3. 集計結果・分析コメント付きの **Excelレポート** をダウンロードできます
""")

uploaded_files = st.file_uploader("📁 ファイルをアップロード（CSVまたはExcel）", type=["csv", "xlsx"], accept_multiple_files=True)

if uploaded_files:
    dfs = []
    for file in uploaded_files:
        try:
            df = pd.read_csv(file) if file.name.endswith(".csv") else pd.read_excel(file)
            dfs.append(df)
        except:
            st.error(f"{file.name} の読み込みに失敗しました。")

    if dfs:
        df_all = pd.concat(dfs, ignore_index=True)
        if "日付" not in df_all.columns:
            st.error("❌ データに「日付」列がありません。1列目に日付を含めてください。")
        else:
            df_all["日付"] = pd.to_datetime(df_all["日付"], errors="coerce")
            df_all = df_all.dropna(subset=["日付"])

            output = BytesIO()
            wb = Workbook()
            wb.remove(wb.active)

            # 指標があれば自動計算
            def calc_metrics(df):
                result = {}
                if "売上" in df.columns and "費用" in df.columns:
                    result["ROAS"] = df["売上"].sum() / df["費用"].sum() if df["費用"].sum() > 0 else 0
                    result["ROI"] = (df["売上"].sum() - df["費用"].sum()) / df["費用"].sum() if df["費用"].sum() > 0 else 0
                if "費用" in df.columns and "コンバージョン" in df.columns:
                    result["CPA"] = df["費用"].sum() / df["コンバージョン"].sum() if df["コンバージョン"].sum() > 0 else 0
                if "売上" in df.columns and "コンバージョン" in df.columns:
                    result["LTV"] = df["売上"].sum() / df["コンバージョン"].sum() if df["コンバージョン"].sum() > 0 else 0
                return result

            # 集計（頻度別）
            for freq, label in [("D", "日別"), ("W", "週別"), ("M", "月別")]:
                grouped = df_all.groupby(pd.Grouper(key="日付", freq=freq)).sum(numeric_only=True)
                grouped.reset_index(inplace=True)
                ws = wb.create_sheet(title=label)
                for row in dataframe_to_rows(grouped, index=False, header=True):
                    ws.append(row)

            # KPIシート
            ws_kpi = wb.create_sheet("KPI")
            kpi_values = calc_metrics(df_all)
            ws_kpi.append(["指標", "値"])
            for k, v in kpi_values.items():
                ws_kpi.append([k, round(v, 2)])

            # 分析コメント
            comment = "このデータには以下の指標が含まれています：

"
            if not kpi_values:
                comment += "- 分析可能な指標が見つかりませんでした。"
            else:
                for k, v in kpi_values.items():
                    if k == "ROAS":
                        comment += f"- ROAS（広告費対効果）は {round(v, 2)} です。
"
                        if v < 1:
                            comment += "  → 広告費に対して売上が低めです。広告のクリエイティブやターゲティングの見直しが必要かもしれません。
"
                    if k == "LTV":
                        comment += f"- LTV（顧客生涯価値）は {round(v, 2)} 円です。
"
                        if v < 1000:
                            comment += "  → 顧客単価が低い傾向です。アップセルやリピート施策の検討が必要です。
"

            ws_cmt = wb.create_sheet("分析コメント")
            for i, line in enumerate(comment.split("\n")):
                ws_cmt.cell(row=i + 1, column=1, value=line)

            wb.save(output)
            st.success("✅ レポートを生成しました。以下からダウンロードできます。")
            st.download_button("📥 Excelレポートをダウンロード", data=output.getvalue(), file_name="マーケティングレポート.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.warning("読み込めるファイルがありません。")
