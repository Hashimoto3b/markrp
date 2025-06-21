import streamlit as st
import pandas as pd
import os
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Webマーケティングレポート生成", layout="centered")

st.title("📊 Webマーケティング自動レポート作成ツール")
st.markdown("CSVまたはExcelファイルをアップロードしてください。自動で集計し、Excelレポートを出力します。")

uploaded_files = st.file_uploader("ファイルを1つ以上アップロード", type=["csv", "xlsx"], accept_multiple_files=True)

if uploaded_files:
    dataframes = []
    for file in uploaded_files:
        try:
            if file.name.endswith(".csv"):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            dataframes.append(df)
        except Exception as e:
            st.error(f"{file.name} の読み込みに失敗しました。")

    if not dataframes:
        st.warning("有効なファイルがありません。")
    else:
        df_all = pd.concat(dataframes, ignore_index=True)

        if "日付" not in df_all.columns:
            st.error("「日付」列が必要です。")
        else:
            df_all["日付"] = pd.to_datetime(df_all["日付"], errors="coerce")
            df_all = df_all.dropna(subset=["日付"])

            output = BytesIO()
            wb = Workbook()
            wb.remove(wb.active)

            for freq, label in [("D", "日別"), ("W", "週別"), ("M", "月別")]:
                grouped = df_all.groupby(pd.Grouper(key="日付", freq=freq)).sum(numeric_only=True)
                grouped.reset_index(inplace=True)
                ws = wb.create_sheet(title=label)
                for row in dataframe_to_rows(grouped, index=False, header=True):
                    ws.append(row)

            ws_comment = wb.create_sheet("分析コメント")
            ws_comment["A1"] = "アップロードされたデータをもとに、日別・週別・月別の集計を行いました。"
            ws_comment["A2"] = "PDF出力は含まれていません。Excelレポートのみ出力されます。"

            wb.save(output)
            st.success("レポートが生成されました！")
            st.download_button(label="📥 Excelレポートをダウンロード", data=output.getvalue(),
                               file_name="web_marketing_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
