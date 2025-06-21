import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title="Webマーケティングレポート", layout="centered")

st.title("📊 Webマーケティング自動レポート作成ツール（最新版）")
st.markdown("アップロードされたデータから、自動で日別・週別・月別の集計、KPI（LTV、ROAS、ROI、CPA）の算出、改善コメントを生成します。")

# 初心者向けガイド
with st.expander("🟢 初めて使う方へ（クリックで開閉）", expanded=True):
    st.markdown("""
1. **CSVまたはExcelファイルをアップロード**してください（複数ファイル可）  
2. 自動で指標（LTV、ROASなど）を計算し、改善ポイントを表示します  
3. Excel形式でレポートをダウンロードできます
    """)

uploaded_files = st.file_uploader("📁 ファイルをアップロード（CSVまたはExcel）", type=["csv", "xlsx"], accept_multiple_files=True)

# 柔軟な列判定
def find_column(possible_names, columns):
    for name in possible_names:
        for col in columns:
            if name in col:
                return col
    return None

if uploaded_files:
    dfs = []
    for file in uploaded_files:
        df = pd.read_csv(file) if file.name.endswith(".csv") else pd.read_excel(file)
        dfs.append(df)

    if dfs:
        df_all = pd.concat(dfs, ignore_index=True)
        df_all.columns = df_all.columns.astype(str)

        # 日付列判定
        date_col = find_column(["日付", "date"], df_all.columns)
        if not date_col:
            st.error("❌ データに「日付」列が見つかりません。")
        else:
            df_all[date_col] = pd.to_datetime(df_all[date_col], errors="coerce")
            df_all = df_all.dropna(subset=[date_col])

            # 指標列を特定
            revenue_col = find_column(["売上", "revenue", "金額"], df_all.columns)
            cost_col = find_column(["費用", "広告費", "cost"], df_all.columns)
            cv_col = find_column(["コンバージョン", "CV", "成約"], df_all.columns)

            # 集計単位
            output = BytesIO()
            wb = Workbook()
            wb.remove(wb.active)

            for freq, label in [("D", "日別"), ("W", "週別"), ("M", "月別")]:
                grouped = df_all.groupby(pd.Grouper(key=date_col, freq=freq)).sum(numeric_only=True)
                grouped.reset_index(inplace=True)
                ws = wb.create_sheet(title=label)
                for row in dataframe_to_rows(grouped, index=False, header=True):
                    ws.append(row)

            # KPI算出
            ws_kpi = wb.create_sheet("KPI")
            ws_kpi.append(["指標", "値"])
            kpi_comments = []
            if revenue_col and cost_col:
                total_revenue = df_all[revenue_col].sum()
                total_cost = df_all[cost_col].sum()
                if total_cost > 0:
                    roas = total_revenue / total_cost
                    roi = (total_revenue - total_cost) / total_cost
                    ws_kpi.append(["ROAS", round(roas, 2)])
                    ws_kpi.append(["ROI", round(roi, 2)])
                    kpi_comments.append(f"- ROASは {round(roas,2)}。{'広告費に対して売上が十分でない可能性があります。' if roas < 1 else '良好なROASです。'}")
                else:
                    ws_kpi.append(["ROAS", "費用がゼロのため算出不可"])

            if revenue_col and cv_col:
                total_cv = df_all[cv_col].sum()
                if total_cv > 0:
                    ltv = total_revenue / total_cv
                    ws_kpi.append(["LTV", round(ltv, 2)])
                    kpi_comments.append(f"- LTVは {round(ltv,2)}。{'単価が低いためアップセル施策が有効です。' if ltv < 1000 else '顧客単価は安定しています。'}")
                else:
                    ws_kpi.append(["LTV", "CV数がゼロのため算出不可"])

            if cost_col and cv_col:
                if total_cv > 0:
                    cpa = total_cost / total_cv
                    ws_kpi.append(["CPA", round(cpa, 2)])
                    kpi_comments.append(f"- CPAは {round(cpa,2)}。{'費用対効果が低い可能性があります。' if cpa > 3000 else '獲得効率は良好です。'}")
                else:
                    ws_kpi.append(["CPA", "CV数がゼロのため算出不可"])

            # コメント出力
            ws_comment = wb.create_sheet("分析コメント")
            ws_comment.append(["改善ポイント・分析結果"])
            for line in kpi_comments:
                ws_comment.append([line])

            wb.save(output)
            st.success("✅ レポートを生成しました。以下からダウンロードできます。")
            st.download_button("📥 Excelレポートをダウンロード", data=output.getvalue(), file_name="自動マーケレポート.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
